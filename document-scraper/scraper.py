# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "trafilatura>=2.0.0",
#   "requests>=2.34.0",
#   "beautifulsoup4>=4.14.3",
#   "lxml>=6.1.0",
#   "openpyxl>=3.1.0",
# ]
# ///
"""
Optilogic ドキュメントスクレイパー
使い方:
  uv run scraper.py          # URL取得 + コンテンツ取得 + Excel変換（フル実行）
  uv run scraper.py --urls   # URLリスト取得のみ → all_urls.txt
  uv run scraper.py --fetch  # コンテンツ取得のみ（all_urls.txt が必要）
  uv run scraper.py --excel  # Excelファイル変換のみ（notebook_source.md に追記）

認証不要: Optilogic のヘルプセンターは公開サイトのため Cookie 不要。
"""

import sys
import time
import requests
import trafilatura
import openpyxl
from bs4 import BeautifulSoup
from pathlib import Path

# --- 設定 ---
SITEMAP_URL = "https://optilogic.com/sitemap.xml"
BASE_URL = "https://optilogic.com"
# 取得対象のURLパターン（help-center 配下のみ）
INCLUDE_PATHS = [
    "/resources/help-center/docs/",
    "/resources/help-center/getting-started",
    "/resources/help-center/navigating",
    "/resources/help-center/connecting",
    "/resources/help-center/troubleshoot",
    "/resources/help-center/knowledge-library",
]
URLS_FILE = Path(__file__).parent / "all_urls.txt"
OUTPUT_FILE = Path(__file__).parent / "notebook_source.md"
REQUEST_DELAY = 1.5  # リクエスト間隔（秒）

# Anura データ構造 Excel ファイル（入力・出力仕様書）
EXCEL_SPECS = [
    {
        "name": "Anura 2.8 Input Data Structure（入力テーブル仕様書）",
        "url": "https://cdn.prod.website-files.com/682c88de65bdb86ec53f8277/69ca874e7096944aaa479a22_Anura-2.8-Inputs.xlsx",
        "filename": "Anura-2.8-Inputs.xlsx",
        "source_page": "https://optilogic.com/resources/help-center/docs/downloadable-anura-data-structure---inputs",
    },
    {
        "name": "Anura 2.8 Output Data Structure（出力テーブル仕様書）",
        "url": "https://cdn.prod.website-files.com/682c88de65bdb86ec53f8277/69ca874e249c4897ef8e26e1_Anura-2.8-Outputs.xlsx",
        "filename": "Anura-2.8-Outputs.xlsx",
        "source_page": "https://optilogic.com/resources/help-center/docs/downloadable-anura-data-structure---outputs",
    },
]
# ------------


def make_headers() -> dict:
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36"
    }


def is_target_url(url: str) -> bool:
    """ヘルプセンター配下のURLかどうかを判定"""
    return any(pattern in url for pattern in INCLUDE_PATHS)


def parse_sitemap_xml(xml_text: str) -> tuple[list[str], list[str]]:
    """XML文字列からサブサイトマップURLとページURLを返す"""
    soup = BeautifulSoup(xml_text, "lxml-xml")
    if soup.find("sitemapindex"):
        sub_sitemaps = [loc.text.strip() for loc in soup.select("sitemap > loc")]
        return sub_sitemaps, []
    page_urls = [loc.text.strip() for loc in soup.select("url > loc")]
    return [], page_urls


def fetch_all_urls_from_sitemap() -> list[str]:
    """サイトマップを再帰的に巡回して全ページURLを取得"""
    headers = make_headers()
    all_urls = []

    print(f"メインサイトマップ取得: {SITEMAP_URL}")
    try:
        r = requests.get(SITEMAP_URL, headers=headers, timeout=15)
        r.raise_for_status()
        sub_sitemaps, page_urls = parse_sitemap_xml(r.text)

        if page_urls:
            all_urls.extend(page_urls)

        for sub_url in sub_sitemaps:
            print(f"  サブサイトマップ: {sub_url}")
            try:
                sr = requests.get(sub_url, headers=headers, timeout=15)
                sr.raise_for_status()
                _, urls = parse_sitemap_xml(sr.text)
                all_urls.extend(urls)
                time.sleep(0.5)
            except Exception as e:
                print(f"  [エラー] {sub_url} → {e}")

    except Exception as e:
        print(f"[エラー] サイトマップ取得失敗: {e}")

    return all_urls


def step_fetch_urls():
    """Step 1: サイトマップからヘルプセンターのURLを取得して all_urls.txt に保存"""
    print("=== Step 1: URLリスト取得 ===")

    all_urls = fetch_all_urls_from_sitemap()

    # ヘルプセンター配下のURLのみ絞り込み
    target_urls = [u for u in all_urls if is_target_url(u)]
    unique_urls = sorted(set(target_urls))

    if not unique_urls:
        print("[失敗] 対象URLが取得できませんでした。")
        return []

    URLS_FILE.write_text("\n".join(unique_urls) + "\n", encoding="utf-8")
    print(f"[OK] {len(unique_urls)} 件のURLを {URLS_FILE.name} に保存しました。\n")
    return unique_urls


def step_fetch_contents(urls: list[str] | None = None):
    """Step 2: URLリストからコンテンツを取得して notebook_source.md に保存"""
    print("=== Step 2: コンテンツ取得 ===")

    if urls is None:
        if not URLS_FILE.exists():
            print(f"[エラー] {URLS_FILE.name} がありません。先に --urls を実行してください。")
            return
        urls = [line.strip() for line in URLS_FILE.read_text(encoding="utf-8").splitlines() if line.strip()]

    headers = make_headers()
    print(f"{len(urls)} 件のURLからコンテンツを取得します...")
    fetched = 0
    errors = 0

    with open(OUTPUT_FILE, "w", encoding="utf-8") as out:
        out.write("# Optilogic ドキュメント - 全コンテンツ\n")
        out.write("# CosmicFrog / DataStar / Ada / Composable Apps 操作ガイド\n\n")

        for i, url in enumerate(urls, 1):
            print(f"[{i}/{len(urls)}] {url}")
            try:
                r = requests.get(url, headers=headers, timeout=20)
                r.raise_for_status()
                content = trafilatura.extract(
                    r.text,
                    include_comments=False,
                    include_tables=True,
                    no_fallback=False,
                )
                if content:
                    # ページタイトルを取得
                    soup = BeautifulSoup(r.text, "lxml")
                    title = soup.find("title")
                    title_text = title.get_text(strip=True) if title else url.split("/")[-1]
                    out.write(f"\n\n---\n## {title_text}\n**URL:** {url}\n\n{content}\n")
                    fetched += 1
                else:
                    print(f"  [スキップ] コンテンツ抽出できず")
                time.sleep(REQUEST_DELAY)
            except Exception as e:
                errors += 1
                print(f"  [エラー] {e}")

    print(f"\n[OK] {fetched}/{len(urls)} 件を {OUTPUT_FILE.name} に保存しました。")
    if errors:
        print(f"[警告] {errors} 件でエラーが発生しました。")
    print("\n次のステップ:")
    print("  1. notebook_source.md を NotebookLM または Claude Project にアップロード")
    print("  2. system_prompt.md の内容を Claude Project の指示に貼り付ける\n")


def excel_sheet_to_markdown(sheet) -> str:
    """openpyxl シートをマークダウンテーブルに変換"""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ""

    # 空行・完全空列を除去
    non_empty_rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
    if not non_empty_rows:
        return ""

    # 最大列数を揃える
    max_cols = max(len(r) for r in non_empty_rows)
    padded = [list(r) + [None] * (max_cols - len(r)) for r in non_empty_rows]

    def cell_str(v) -> str:
        if v is None:
            return ""
        return str(v).replace("|", "｜").replace("\n", " ").strip()

    # ヘッダー行
    header = padded[0]
    md = "| " + " | ".join(cell_str(c) for c in header) + " |\n"
    md += "| " + " | ".join("---" for _ in header) + " |\n"
    for row in padded[1:]:
        md += "| " + " | ".join(cell_str(c) for c in row) + " |\n"
    return md


def step_fetch_excel():
    """Step 3: Anura Excel仕様書をダウンロードしてマークダウンに変換し notebook_source.md に追記"""
    print("=== Step 3: Excel仕様書取得・変換 ===")

    headers = make_headers()

    with open(OUTPUT_FILE, "a", encoding="utf-8") as out:
        for spec in EXCEL_SPECS:
            print(f"ダウンロード: {spec['name']}")
            local_path = Path(__file__).parent / spec["filename"]
            try:
                r = requests.get(spec["url"], headers=headers, timeout=60)
                r.raise_for_status()
                local_path.write_bytes(r.content)
                print(f"  保存完了: {local_path.name} ({len(r.content):,} bytes)")

                wb = openpyxl.load_workbook(local_path, data_only=True)
                out.write(f"\n\n---\n## {spec['name']}\n")
                out.write(f"**ソースURL:** {spec['source_page']}\n")
                out.write(f"**ファイル:** {spec['filename']}\n\n")
                out.write(
                    "このファイルは Cosmic Frog (Anura エンジン) の入出力テーブル仕様書です。\n"
                    "各シートにテーブル名・フィールド名・データ型・説明が記載されています。\n\n"
                )

                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    md_table = excel_sheet_to_markdown(ws)
                    if md_table:
                        out.write(f"### シート: {sheet_name}\n\n")
                        out.write(md_table + "\n")

                wb.close()
                print(f"  変換完了: {len(wb.sheetnames)} シートを notebook_source.md に追記")

            except Exception as e:
                print(f"  [エラー] {e}")

    print("\n[OK] Excel仕様書の変換が完了しました。\n")


def main():
    args = sys.argv[1:]

    if "--urls" in args:
        step_fetch_urls()
    elif "--fetch" in args:
        step_fetch_contents()
    elif "--excel" in args:
        step_fetch_excel()
    else:
        urls = step_fetch_urls()
        if urls:
            step_fetch_contents(urls)
        step_fetch_excel()


if __name__ == "__main__":
    main()
